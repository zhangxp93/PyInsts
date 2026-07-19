import os
import sqlite3
import logging
import pandas as pd
from datetime import datetime
from typing import Optional, List, Dict, Any
from contextlib import contextmanager
from queue import Queue
from threading import Lock
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

class SQLiteHandler:
    """SQLite数据库操作类"""
    def __init__(self, sql_db_path: str = 'test_data.db', export_excel: bool = True):
        self.db_path = sql_db_path
        self._ensure_db_directory()
        self._pool = self._create_connection_pool()
        self._create_metadata_table()
        self._tables_to_export = set()  # 使用集合存储需要导出的表名
        self._data_cache = {}  # 缓存数据，用于最后一次性导出
        self._total_rows = 0  # 记录总行数
        self._export_excel = export_excel  # 控制是否导出Excel

    def _create_connection_pool(self) -> Queue:
        """创建连接池"""
        pool = Queue(maxsize=5)
        for _ in range(5):
            conn = sqlite3.connect(self.db_path, timeout=30)    # 创建连接池
            pool.put(conn)
        return pool

    def _ensure_db_directory(self):
        """确保数据库目录存在"""
        os.makedirs(os.path.dirname(self.db_path), exist_ok=True)

    def _create_metadata_table(self):
        """创建元数据表"""
        with self._get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS metadata (
                    table_name TEXT PRIMARY KEY,
                    created_at TIMESTAMP,
                    last_updated TIMESTAMP,
                    row_count INTEGER
                )
            """)
            conn.commit()

    @contextmanager
    def _get_connection(self):
        """获取数据库连接"""
        conn = self._pool.get()
        try:
            yield conn
        finally:
            self._pool.put(conn)

    def create_table_if_not_exists(self, table_name: str, df: pd.DataFrame) -> None:
        """如果表不存在，则创建。字段名基于DataFrame的列"""
        with self._get_connection() as conn:
            cursor = conn.cursor()

            # 检查是否有图片数据（支持更多图片相关列名）
            image_columns = [col for col in df.columns if '图片' in col or 'image' in col.lower()]
            has_image_data = any(df[col].notna().any() and df[col].astype(str).str.strip().str.len().gt(0).any()
                               for col in image_columns)

            # 获取列类型
            column_types = self._infer_column_types(df)

            # 构建列定义
            columns = []
            for col, dtype in column_types.items():
                if col not in image_columns:  # 排除图片相关列
                    columns.append(f"[{col}] {dtype}")
                else:  # 保留原始图片列名
                    columns.append(f"[{col}] TEXT")

            # 创建表
            sql = f"""
                CREATE TABLE IF NOT EXISTS [{table_name}] (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    {', '.join(columns)}
                )
            """
            cursor.execute(sql)

            # 更新元数据
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            cursor.execute("""
                INSERT OR IGNORE INTO metadata 
                (table_name, created_at, last_updated, row_count)
                VALUES (?, ?, ?, 0)
            """, (table_name, current_time, current_time))

            conn.commit()

    def _infer_column_types(self, df: pd.DataFrame) -> Dict[str, str]:
        """推断列的数据类型"""
        type_mapping = {
            'int64': 'INTEGER',
            'float64': 'REAL',
            'bool': 'INTEGER',
            'datetime64[ns]': 'TIMESTAMP',
            'object': 'TEXT'
        }
        return {col: type_mapping.get(str(df[col].dtype), 'TEXT')
                for col in df.columns}

    def _validate_table_name(self, table_name: str) -> str:
        """验证并清理表名，防止SQL注入"""
        # 只允许字母、数字、下划线和中文
        import re
        if not table_name:
            raise ValueError("表名不能为空")
        # 移除特殊字符（根据需要调整）
        clean_name = re.sub(r'[^\w\u4e00-\u9fa5]', '', table_name)
        if not clean_name:
             raise ValueError(f"无效的表名: {table_name}")
        return clean_name

    def insert_data(self, table_name: str, df: pd.DataFrame) -> None:
        """
        将DataFrame插入数据库 (优化版：批量插入)
        """
        if df.empty:
            return

        self.create_table_if_not_exists(table_name, df)

        with self._get_connection() as conn:
            cursor = conn.cursor()

            # 检查是否有图片数据（支持更多图片相关列名）
            image_columns = [col for col in df.columns if '图片' in col or 'image' in col.lower()]
            has_image_data = any(df[col].notna().any() and df[col].astype(str).str.strip().str.len().gt(0).any()
                               for col in image_columns)

            # 获取普通数据列名
            data_columns = [col for col in df.columns if col not in image_columns]
            
            # 准备插入的列名列表，保留原始图片列名
            insert_columns = [f'[{col}]' for col in df.columns]
            
            placeholders = ', '.join(['?'] * len(insert_columns))
            insert_sql = f"""
                INSERT INTO [{table_name}] 
                ({', '.join(insert_columns)}) 
                VALUES ({placeholders})
            """

            # 准备批量插入的数据
            batch_values = []
            
            try:
                # 预处理数据
                for _, row in df.iterrows():
                    row_values = []
                    # 处理所有列的数据
                    for col in df.columns:
                        if col in image_columns:
                            # 处理图片列
                            image_path = row[col] if col in row and pd.notna(row[col]) else None
                            row_values.append(image_path)
                        else:
                            # 处理非图片列
                            val = row[col]
                            # 将numpy类型转换为原生python类型，避免sqlite适配问题
                            if hasattr(val, 'item'): 
                                 val = val.item()
                            row_values.append(val)
                    
                    batch_values.append(tuple(row_values))

                # 执行批量插入
                if batch_values:
                    cursor.executemany(insert_sql, batch_values)
                    
                    # 更新元数据
                    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    cursor.execute("""
                        UPDATE metadata 
                        SET last_updated = ?, 
                            row_count = row_count + ? 
                        WHERE table_name = ?
                    """, (current_time, len(batch_values), table_name))

                    conn.commit()

                    # 记录需要导出的表名并缓存数据
                    self._tables_to_export.add(table_name)
                    # 注意：缓存可能会占用大量内存，如果数据量特别大建议根据需求调整
                    if table_name not in self._data_cache:
                        self._data_cache[table_name] = []
                    self._data_cache[table_name].append(df)
                    self._total_rows += len(df)
                    
                    logger.info(f"成功批量插入 {len(batch_values)} 条数据到表 {table_name}")

            except sqlite3.Error as e:
                logger.error("SQLite错误: %s", str(e))
                conn.rollback()
                raise
            except Exception as e:
                logger.error("插入数据错误: %s", str(e))
                conn.rollback()
                raise

    def insert_data_list(self, table_name: str, data_list: List) -> None:
        """
        将DataFrame列表合并后插入数据库

        Args:
            table_name: 表名
            data_list: 数据列表
        """
        # 一次性创建DataFrame
        df = pd.DataFrame(data_list)
        data_list = [df]

        # 合并所有数据
        all_data = pd.concat(data_list, ignore_index=True)
        # 插入数据
        self.insert_data(table_name, all_data)

    def export_to_excel(self, excel_path: str = None) -> None:
        """
        将数据库中的所有表导出为Excel文件，包含图片路径超链接

        Args:
            excel_path: Excel文件路径，如果为None则使用数据库路径
        """
        if excel_path is None:
            excel_path = os.path.splitext(self.db_path)[0] + '.xlsx'

        try:
            # 创建ExcelWriter
            writer = pd.ExcelWriter(excel_path, engine='openpyxl')

            # 获取数据库中的所有表
            with self._get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT name FROM sqlite_master 
                    WHERE type='table' AND name NOT LIKE 'sqlite_%' AND name != 'metadata'
                """)
                tables = cursor.fetchall()

                # 导出每个表
                for table in tables:
                    table_name = table[0]
                    
                    # 过滤掉不需要导出的表 (例如以 _JSON 结尾的)
                    if table_name.endswith('_JSON'):
                        logger.info(f"跳过导出 JSON 表: {table_name}")
                        continue

                    # 获取数据，包括图片数据
                    df = pd.read_sql_query(f"""
                        SELECT * FROM [{table_name}] 
                        ORDER BY id
                    """, conn)

                    # 移除id列
                    df.drop(columns=['id'], inplace=True, errors='ignore')

                    # 检查是否存在图片相关列
                    image_columns = [col for col in df.columns if '图片' in col or 'image' in col.lower()]
                    existing_image_columns = [col for col in image_columns if col in df.columns]

                    # 如果存在图片列，检查是否有实际的图片数据
                    has_image_data = any(df[col].notna().any() and df[col].astype(str).str.strip().str.len().gt(0).any()
                                       for col in image_columns)

                    # 如果没有实际的图片数据，移除图片相关列
                    if not has_image_data and existing_image_columns:
                        df.drop(columns=existing_image_columns, inplace=True, errors='ignore')
                        logger.info(f"移除空的图片列: {existing_image_columns}")

                    # 导出到Excel
                    df.to_excel(writer, index=False, sheet_name=table_name)

                    # 获取工作表
                    worksheet = writer.sheets[table_name]

                    for idx, col in enumerate(df.columns):
                        try:
                            max_length = max(
                                df[col].astype(str).apply(len).max(),
                                len(str(col))
                            )
                            # 限制最大列宽为15
                            worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max_length + 2, 15)
                        except Exception as e:
                            logger.warning(f"调整列宽失败 {col}: {str(e)}")
                            worksheet.column_dimensions[get_column_letter(idx + 1)].width = 15  # 设置默认宽度

                        # 设置表头行高为智能行高，最大50
                    header_height = max(len(str(col)) * 3 for col in df.columns)  # 根据列名长度计算行高
                    worksheet.row_dimensions[1].height = min(header_height, 50)  # 限制最大行高为50
                    
                    # 设置表头自动换行和垂直居中
                    for col_idx, col in enumerate(df.columns):
                        header_cell = worksheet.cell(row=1, column=col_idx + 1)
                        header_cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

                    # 如果有图片数据，添加超链接
                    if has_image_data:
                        for idx, row in df.iterrows():
                            # 处理所有图片列
                            for col in image_columns:
                                # 获取图片路径
                                image_path = row.get(col, '')
                                if image_path and os.path.exists(image_path):
                                    try:
                                        # 获取图片路径列的索引
                                        path_col_idx = df.columns.get_loc(col)
                                        # 创建超链接
                                        cell = worksheet.cell(row=idx + 2, column=path_col_idx + 1)
                                        cell.hyperlink = image_path
                                        cell.style = "Hyperlink"
                                    except Exception as e:
                                        logger.error(f"添加超链接失败: {str(e)}")

                    logger.info("成功导出表 %s 到 Excel：%s", table_name, excel_path)

            # 保存Excel文件
            writer.close()

        except Exception as e:
            logger.error("导出Excel失败: %s", str(e))
            raise

    def _export_to_excel(self) -> None:
        """自动导出Excel（内部使用）"""
        if not self._export_excel:
            return
        self.export_to_excel()

    def close(self) -> None:
        """关闭所有连接并导出Excel"""
        # 导出所有表到Excel
        self._export_to_excel()

        # 清空缓存
        self._tables_to_export.clear()
        self._data_cache.clear()
        self._total_rows = 0

        # 关闭所有连接
        while not self._pool.empty():
            conn = self._pool.get()
            conn.close()

    def __enter__(self):
        """上下文管理器入口"""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """上下文管理器出口"""
        self.close()

    def get_image(self, table_name: str, row_id: int, column_name: str = '图片') -> Optional[str]:
        """
        从数据库中获取指定行的图片路径

        Args:
            table_name: 表名
            row_id: 行ID
            column_name: 列名，默认为'图片'

        Returns:
            图片路径，如果不存在则返回None
        """
        with self._get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(f"""
                SELECT [{column_name}] FROM [{table_name}] 
                WHERE id = ?
            """, (row_id,))
            result = cursor.fetchone()
            return result[0] if result else None

    def export_image(self, table_name: str, row_id: int, output_path: str) -> bool:
        """
        此方法已弃用，因为数据库中只保存图片链接地址

        Args:
            table_name: 表名
            row_id: 行ID
            output_path: 输出路径

        Returns:
            始终返回False
        """
        logger.warning("此方法已弃用，因为数据库中只保存图片链接地址")
        return False

    def view_image(self, table_name: str, row_id: int) -> None:
        """
        此方法已弃用，因为数据库中只保存图片链接地址

        Args:
            table_name: 表名
            row_id: 行ID
        """
        logger.warning("此方法已弃用，因为数据库中只保存图片链接地址")

    def import_from_excel(self, excel_path: str, table_name: str = None) -> None:
        """
        从Excel文件导入数据到数据库

        Args:
            excel_path: Excel文件路径
            table_name: 目标表名，如果为None则使用Excel文件名作为表名
        """
        try:
            # 读取Excel文件的所有工作表
            logger.info(f"开始读取Excel文件: {excel_path}")
            excel_file = pd.ExcelFile(excel_path)
            sheet_names = excel_file.sheet_names

            for sheet_name in sheet_names:
                # 如果没有指定表名，使用Excel文件名和工作表名组合作为表名
                if table_name is None:
                    current_table_name = f"{os.path.splitext(os.path.basename(excel_path))[0]}_{sheet_name}"
                else:
                    current_table_name = f"{table_name}_{sheet_name}"

                # 读取当前工作表
                df = pd.read_excel(excel_path, sheet_name=sheet_name)

                # 处理时间戳类型
                for col in df.columns:
                    if pd.api.types.is_datetime64_any_dtype(df[col]):
                        # 将时间戳转换为字符串格式
                        df[col] = df[col].dt.strftime('%Y-%m-%d %H:%M:%S')
                    elif pd.api.types.is_timedelta64_dtype(df[col]):
                        # 将时间差转换为字符串格式
                        df[col] = df[col].astype(str)

                # 检查是否有图片相关列
                image_columns = [col for col in df.columns if '图片' in col or 'image' in col.lower()]

                # 处理图片数据
                if image_columns:
                    for col in image_columns:
                        # 仅保留图片路径，不验证文件是否存在
                        df[col] = df[col].apply(lambda x: x if pd.notna(x) else None)

                # 插入数据到数据库
                logger.info(f"开始导入数据到表: {current_table_name}")
                self.insert_data(current_table_name, df)

                logger.info(f"成功从Excel文件导入数据到表 {current_table_name}")

        except Exception as e:
            logger.error(f"从Excel导入数据失败: {str(e)}")
            raise


    def save_data(self, table_name: str, data: Dict[str, Any]) -> None:
        """
        处理并保存各种格式的数据（自动展开列表）

        Args:
            table_name: 表名
            data: 数据字典
        """
        all_data = []
        # 使用 DataProcessor 处理数据
        processed_data = DataProcessor.process_data(data)
        all_data.extend(processed_data)
        
        # 批量插入
        self.insert_data_list(table_name=table_name, data_list=all_data)


class ExcelHandler:
    """Excel处理类"""
    def __init__(self):
        self.file_path: Optional[str] = None
        self.workbook: Optional[Workbook] = None
        self._lock = Lock()

    def create_excel(self, excel_path: str = "data.xlsx",
                    title: Optional[str] = None) -> None:
        """
        创建或打开Excel文件

        Args:
            excel_path: Excel文件路径
            title: 工作表名称
        """
        self.file_path = excel_path
        if not title or title.strip() == '':
            title = 'Sheet1'
            logger.info("未指定工作表名称，使用默认名称 '%s'", title)

        try:
            # 确保目录存在
            os.makedirs(os.path.dirname(excel_path), exist_ok=True)

            # 尝试加载现有文件
            try:
                self.workbook = load_workbook(excel_path)
                if title in self.workbook.sheetnames:
                    logger.info("工作表 %s 已存在，追加数据...", title)
                else:
                    self.workbook.create_sheet(title)
            except FileNotFoundError:
                # 文件不存在，创建新文件
                self.workbook = Workbook()
                sheet = self.workbook.active
                sheet.title = title
                logger.info("文件 %s 不存在，已创建新文件和新工作表...", excel_path)
            except InvalidFileException:
                raise ValueError(f"文件格式无效: {excel_path}")
            except Exception as e:
                raise RuntimeError(f"处理Excel文件时发生错误: {str(e)}")

        except Exception as e:
            logger.error("创建Excel文件失败: %s", str(e))
            raise

    def get_sheet(self, title: str) -> Any:
        """
        获取指定名称的工作表

        Args:
            title: 工作表名称

        Returns:
            工作表对象
        """
        if not self.workbook:
            raise RuntimeError("工作簿未初始化，请先调用create_excel")

        if title not in self.workbook.sheetnames:
            raise ValueError(f"工作表 '{title}' 不存在")

        return self.workbook[title]

    def save_excel(self) -> None:
        """保存Excel文件"""
        if not self.workbook or not self.file_path:
            raise RuntimeError("工作簿或文件路径未初始化")

        try:
            with self._lock:
                self.workbook.save(self.file_path)
                self.workbook.close()
                self.workbook = None
                logger.info("成功保存Excel文件: %s", self.file_path)
        except Exception as e:
            logger.error("保存Excel文件失败: %s", str(e))
            raise

    def __enter__(self):
        """上下文管理器入口"""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """上下文管理器出口"""
        if self.workbook:
            self.save_excel()


class DataProcessor:
    """数据处理类"""
    @staticmethod
    def process_data(data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        处理数据，将列表字段展开成多条记录

        Args:
            data: 原始数据字典

        Returns:
            处理后的数据列表
        """
        # 找出当前数据中所有列表字段的最大长度
        list_values = [v for v in data.values() if isinstance(v, list)]
        if not list_values:
            max_length = 1
        else:
            # 过滤掉空列表，避免 max() 报错（虽然理论上如果有列表，长度至少为0）
            non_empty_lists = [v for v in list_values if len(v) > 0]
            if not non_empty_lists:
                 # 如果全是空列表，则视为空数据，但为了防止报错，设为0或1视具体需求。
                 # 这里假设如果是空列表，则不生成数据或生成一行空数据？
                 # 根据上下文，如果有关联键值对，非列表项应该保留。
                 # 我们取最大长度为1（保留非列表项）或者0（如果真的全是空）。
                 # 安全起见，如果存在非列表项，至少产生1行。
                 max_length = 1
            else:
                 max_length = max(len(v) for v in non_empty_lists)

        # 将非列表值转换为相同长度的列表
        for key, value in data.items():
            if not isinstance(value, list):
                data[key] = [value] * max_length
            elif len(value) < max_length:
                if len(value) == 0:
                     # 处理空列表的情况：填充 None
                     data[key] = [None] * max_length
                else:
                     # 如果列表长度不足，用最后一个值填充
                     data[key] = value + [value[-1]] * (max_length - len(value))

        # 创建所有可能的组合
        processed_data = []
        for j in range(max_length):
            record = {key: value[j] for key, value in data.items()}
            processed_data.append(record)

        return processed_data


def save_data_to_spl(db_path,table_name,data,export_excel=False):
    with SQLiteHandler(db_path, export_excel=export_excel) as db:
        # 使用新方法一次性写入所有数据
        # 使用数据处理类处理数据
        db.save_data(table_name, data)


if __name__ == '__main__':
    # 测试数据

    # 使用SQLite存储并自动导出Excel
    db_path = "data/2_test_optimized.db"
    table_name = "eng_test"

    # 【推荐用法】在循环外部打开数据库连接，既快又安全
    # export_excel=False 表示测试过程中不导出Excel，等所有测试做完可以在 finally 块或者单独导出
    with SQLiteHandler(db_path, export_excel=False) as db:
        for i in range(1000):
            current_time = datetime.now()
            data = {
                '测试时间': current_time.strftime('%Y-%m-%d %H:%M:%S'),
                '编号': i,
                '型号': 'demo',  # 这里可以根据需要动态生成
                '测试温度': '+85',  # 这里可以根据需要动态生成
                '设置频率(MHz)': 1000,  # 这里可以根据需要动态生成
                'vt电压': [1.8, 1.9, 2.0],
                '频率(GHz)': [1000, 1100, 1200],  # 这里可以根据需要动态生成
                '输出功率(dBm)': 10,  # 这里可以根据需要动态生成
                '谐波频率': 2000,  # 这里可以根据需要动态生成
                '谐波抑制': 10,  # 这里可以根据需要动态生成
                '-100M鉴相杂散': 10,  # 这里可以根据需要动态生成
                '+100M鉴相杂散': 10,  # 这里可以根据需要动态生成
                '小数杂散频偏(KHz)': 100,  # 这里可以根据需要动态生成
                '小数杂散': 10,  # 这里可以根据需要动态生成
                'vcc5电流': 10,  # 这里可以根据需要动态生成
            }

            # 直接写入！
            # 使用实例方法，更加面向对象
            db.save_data(table_name, data)

            # 简单的进度打印，避免刷屏
            if i % 100 == 0:
                print(f"已保存 {i} 条...")

    print("测试完成，数据已全部保存。")
