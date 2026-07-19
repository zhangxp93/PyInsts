import os
import tempfile
import shutil
import time
import logging

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException




class ExcelHandler:
    """
    创建excel
    """

    def __init__(self):
        self.file_path = None
        self.workbook = None

    def create_excel(self, excel_path="data\\data.xlsx", title=None):
        """
        创建excel
        :param excel_path: excel路径
        :param title: 表名，如果为None则使用默认名称'Sheet1'
        :return:
        """
        self.file_path = excel_path

        # 设置默认工作表名称
        if title is None or title.strip() == '':
            title = 'Sheet1'
            logging.info("未指定工作表名称，使用默认名称 'Sheet1'")

        try:
            self.workbook = load_workbook(self.file_path)
            if title in self.workbook.sheetnames:
                sheet = self.workbook[title]
                logging.info(f"工作表 {title} 已存在，追加数据...")
            else:
                sheet = self.workbook.create_sheet(title)
                logging.info(
                    f"工作表 {title} 不存在，已创建新的工作表并写入数据...")
        except FileNotFoundError:
            self.workbook = Workbook()
            sheet = self.workbook.active
            sheet.title = title
            logging.info(
                f"文件 {excel_path} 不存在，已创建新文件和新工作表并写入数据...")
        except InvalidFileException:
            error_msg = "文件格式无效，请检查文件是否为正确的Excel格式。"
            logging.error(error_msg)
            raise
        except Exception as e:
            error_msg = f"未知错误：{e}"
            logging.error(error_msg)
            raise

    def get_sheet(self, title):
        return self.workbook[title]

    def save_excel(self):
        if self.workbook and self.file_path:
            try:
                temp_dir = os.path.join(os.path.dirname(self.file_path),
                                        '缓存')
                os.makedirs(temp_dir, exist_ok=True)
                temp_file = tempfile.NamedTemporaryFile(delete=False,
                                                        dir=temp_dir)
                self.workbook.save(temp_file.name)
                temp_file.close()

                # 添加文件锁机制
                max_retries = 3
                retry_count = 0
                while retry_count < max_retries:
                    try:
                        shutil.move(temp_file.name, self.file_path)
                        logging.info("文件已成功保存。")
                        break
                    except PermissionError:
                        retry_count += 1
                        logging.warning(
                            f"文件被占用，重试中... ({retry_count}/{max_retries})")
                        time.sleep(1)
                else:
                    raise PermissionError(
                        "文件保存失败：文件可能被占用或没有写入权限。")

                self.workbook.close()
                self.workbook = None
            except Exception as e:
                logging.error(f"保存文件时出错：{e}")
                raise
        else:
            logging.warning("没有工作簿或文件路径，无法保存文件。")

    def close_excel(self):
        if self.workbook:
            try:
                self.workbook.close()
                logging.info("文件已保存并关闭。")
            except Exception as e:
                logging.error(f"关闭文件时出错：{e}")
        else:
            logging.warning("没有工作簿需要关闭。")

    def _get_or_create_sheet(self, sheet_title):
        """获取或创建工作表"""
        if sheet_title is None:
            return self.workbook.active

        if sheet_title not in self.workbook.sheetnames:
            logging.info(f"工作表 {sheet_title} 不存在，正在创建...")
            return self.workbook.create_sheet(sheet_title)
        return self.workbook[sheet_title]

    def _validate_sheet_capacity(self, sheet, data):
        """验证工作表容量"""
        if sheet.max_row + len(data) > 1048576:
            raise ValueError(
                f"工作表 {sheet.title} 剩余容量不足，无法写入 {len(data)} 条数据")

    def save_data(self, data, sheet_title=None):
        """保存数据到Excel
        :param data: 要保存的数据，必须是 pandas DataFrame
        :param sheet_title: 工作表名称，如果为None则使用第一个工作表
        """
        if not isinstance(data, pd.DataFrame):
            raise TypeError("data 参数必须是 pandas DataFrame")

        try:
            sheet = self._get_or_create_sheet(sheet_title)
            self._validate_sheet_capacity(sheet, data)

            # 如果是空表，先写入表头
            if sheet.max_row == 1:
                sheet.append(list(data.columns))

            # 批量写入数据
            data_values = data.values.tolist()
            for row in data_values:
                sheet.append(row)

            logging.info(
                f"成功写入 {len(data)} 条数据到工作表 {sheet.title}")
        except ValueError as ve:
            logging.error(f"数据验证失败：{ve}")
            raise
        except Exception as e:
            logging.error(f"写入数据失败：{e}")
            raise


    def save_data_to_excel(self, excel_path, sheet_title,data_list):
        try:
            self.create_excel(excel_path=excel_path, title=sheet_title)
            for data in data_list:
                self.save_data(data=data, sheet_title=sheet_title)
            self.save_excel()  # 保存整个工作簿
            return True
        except Exception as e:
            logging.error(f"保存数据到Excel时出错：{e}")
            return False


if __name__ == '__main__':
    from datetime import datetime
    data_list = []
    for i in range(10):
        # 准备数据
        current_time = datetime.now()
        data = pd.DataFrame({
            '测试时间': current_time,
            '型号': 'demo',
            '编号': i,
            '测试温度': '+85',
            '设置频率(MHz)': 1000,
            '频率(GHz)': [1000,2000],
            '输出功率(dBm)': 10,
            '谐波频率': 2000,
            '谐波抑制': 10,
            '-100M鉴相杂散': 10,
            '+100M鉴相杂散': 10,
            '小数杂散频偏(KHz)': 100,
            '小数杂散': 10,
            'vcc5电流': 10,
        })
        data_list.append(data)
    excel = ExcelHandler()
    excel_path = "data/data.xlsx"
    sheet_title = 'eng'
    data = data_list
    excel.save_data_to_excel(excel_path=excel_path, sheet_title=sheet_title, data_list=data)